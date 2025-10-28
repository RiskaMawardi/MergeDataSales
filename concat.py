import warnings
warnings.simplefilter("ignore", UserWarning)

import pandas as pd
import os
import glob
import re
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# ====================== KONFIGURASI ======================
folder_path = r'D:\DataFromPrincipal\DataEBLO\merge'
output_file = os.path.join(folder_path, 'EBLO.xlsx')

headers = {
    "SHOPEE": [
        "NO", "BRAND", "No. Pesanan", "Status Pesanan", "Status Pembatalan/ Pengembalian",
        "No. Resi", "Opsi Pengiriman", "Antar ke counter/ pick-up",
        "Pesanan Harus Dikirimkan Sebelum (Menghindari keterlambatan)",
        "Waktu Pengiriman Diatur", "Waktu Pesanan Dibuat", "Waktu Pembayaran Dilakukan",
        "SKU Induk", "Nama Produk", "Nomor Referensi SKU", "Nama Variasi",
        "Harga Sebelum Diskon", "Harga Setelah Diskon", "Jumlah", "Total Harga Produk",
        "Total Diskon", "Diskon Dari Penjual", "Diskon Dari Shopee", "Berat Produk",
        "Jumlah Produk di Pesan", "Total Berat", "Voucher Ditanggung Penjual",
        "Cashback Koin", "Voucher Ditanggung Shopee", "Paket Diskon",
        "Paket Diskon (Diskon dari Shopee)", "Paket Diskon (Diskon dari Penjual)",
        "Potongan Koin Shopee", "Diskon Kartu Kredit", "Ongkos Kirim Dibayar oleh Pembeli",
        "Estimasi Potongan Biaya Pengiriman", "Ongkos Kirim Pengembalian Barang",
        "Total Pembayaran", "Perkiraan Ongkos Kirim", "Catatan dari Pembeli", "Catatan",
        "Username (Pembeli)", "Nama Penerima", "No. Telepon", "Alamat Pengiriman",
        "Kota/Kabupaten", "Provinsi", "Waktu Pesanan Selesai"
    ],
    "TOKPED": [
        "Count", "BRAND", "Invoice", "Payment Date", "Order Status", "Product ID", "Product Name",
        "Quantity", "Stock Keeping Unit (SKU)", "Notes", "Price (Rp)", "Discount Amount (Rp)",
        "Subsidi Amount (Rp)", "Harga Jual (Rp)", "Customer Name", "Customer Phone", "Recipient",
        "Recipient Number", "Recipient Address", "Courier", "Shipping Price + fee (Rp)",
        "Insurance (Rp)", "Total Shipping Fee (Rp)", "Total Amount (Rp)", "AWB", "Jenis Layanan",
        "Bebas Ongkir", "Warehouse Origin", "Campaign Name"
    ],
    "TIKTOK": [
        "NO", "Nama Brand", "Order ID", "Tracking ID", "Cancellation Request", "Product Name",
        "Seller SKU", "Variation", "Quantity", "Paid Time", "Delivery Option", "Buyer Message",
        "Buyer Username", "Recipient", "Phone #", "Zipcode", "Country", "Province",
        "Regency and City", "Districts", "Villages", "Detail Address",
        "Unit Price", "Order Amount", "Payment Method", "Weight(kg)",
        "Product Category", "Purchase Channel"
    ],
    "LAZADA": [
        "NO", "Nama Brand", "Order Item Id", "Lazada Id", "Seller SKU", "Lazada SKU", "Created at",
        "Updated at", "Order Number", "Invoice Required", "Customer Name", "Customer Email",
        "National Registration Number", "Shipping Name", "Shipping Address", "Shipping Address2",
        "Shipping Address3", "Shipping Address4", "Shipping Address5", "Shipping Phone Number",
        "Shipping Phone Number2", "Shipping City", "Shipping Postcode", "Shipping Country",
        "Shipping Region", "Billing Name", "Billing Address", "Billing Address2",
        "Billing Address3", "Billing Address4", "Billing Address5", "Billing Phone Number",
        "Billing Phone Number2", "Billing City", "Billing Country", "Payment Method", "Paid Price",
        "Unit Price", "Shipping Fee", "Wallet Credits", "Item Name", "Variation",
        "CD Shipping Provider", "Shipping Provider", "Shipment Type Name",
        "Shipping Provider Type", "CD Tracking Code", "Tracking Code", "Tracking URL",
        "Shipping Provider (first mile)", "Tracking Code (first mile)", "Tracking URL (first mile)",
        "Promised shipping time", "Premium", "Status", "Cancel / Return Initiator", "Reason",
        "Reason Detail", "Editor", "Bundle ID", "Bundle Discount", "Refund Amount"
    ],
    "BLIBLI": [
        "NO", "Nama Brand", "No. Order", "No. Order Item", "No. Paket", "No. Awb",
        "Tanggal Order", "Nama Pemesan", "No. Tlp", "Kode SKU Blibli", "SKU Merchant",
        "SKU", "Nama Produk", "Total Barang", "Servis Logistik", "Kode Merchant",
        "Order Status", "Alamat", "Kota", "Provinsi", "Harga item pesanan",
        "Total harga item pesanan", "Total", "Catatan produk", "Tanggal pengiriman"
    ]
}

# ====================== UTILITAS CLEANING ======================
def clean_text_general(text: str) -> str:
    if pd.isna(text): return ""
    text = re.sub(r"[^a-zA-Z0-9\s.,*]", "", str(text))
    return re.sub(r"\s+", " ", text).strip()

def clean_numeric(col_series):
    return (
        col_series.astype(str)
        .str.replace(r"[^0-9]", "", regex=True)
        .replace("", "0")
        .astype(float)
    )

def fix_brand(x):
    if pd.isna(x): return ""
    s = str(x).upper()
    s = re.sub(r"^(TIKTOK|TOKOPEDIA)\s+", "", s)
    mapping = {"DR TEAL'S": "DR TEALS"}
    for key, val in mapping.items():
        if key in s:
            s = val
    return s

def parse_datetime_safe(x):
    if pd.isna(x) or str(x).strip() == "":
        return ""
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M"):
        try:
            return pd.to_datetime(x, format=fmt)
        except ValueError:
            continue
    return pd.NaT

# ====================== MERGE + CLEAN ======================
files = glob.glob(os.path.join(folder_path, "*.xlsx"))
combined_sheets = {}

for file_path in files:
    if os.path.basename(file_path).lower() in ["eblo.xlsx", "eblo_clean.xlsx", "eblo_cleaned.xlsx"]:
        continue

    brand = os.path.splitext(os.path.basename(file_path))[0].split(" ")[0]
    print(f"\nProcessing: {brand}")
    xls = pd.ExcelFile(file_path)

    for sheet_name in xls.sheet_names:
        clean_name = sheet_name.strip().upper()
        
        # 🧩 Perbaikan mapping nama sheet
        if clean_name in ["TOKPED NEW", "TOKOPEDIA NEW", "TOKOPEDIA", "TIKTOK"]:
            clean_name = "TIKTOK"
        elif clean_name in ["LAZADAA", "LAZADA ", "LAZADA"]:
            clean_name = "LAZADA"
        elif clean_name in ["BLIBLI", "BLIBLII", "BLIBLI ", "BLIBLIII"]:
            clean_name = "BLIBLI"

        try:
            df = pd.read_excel(xls, sheet_name=sheet_name, dtype=str, header=None)
            df.dropna(how="all", inplace=True)
            if df.empty: continue

            header_row = df.notna().sum(axis=1).idxmax()
            df.columns = df.iloc[header_row].astype(str).str.strip()
            df = df.iloc[header_row + 1:].copy()
            df.dropna(axis=1, how="all", inplace=True)
            df.dropna(how="all", inplace=True)

            # --- Filter invalid rows ---
            key_cols = {
                "SHOPEE": "No. Pesanan",
                "TOKPED": "Invoice",
                "TIKTOK": "Order ID",
                "LAZADA": "Order Item Id",
                "BLIBLI": "No. Order"
            }
            if clean_name in key_cols and key_cols[clean_name] in df.columns:
                key_col = key_cols[clean_name]
                before = len(df)
                df = df[df[key_col].notna() & (df[key_col].astype(str).str.strip() != "")]
                after = len(df)
                if before != after:
                    print(f"➡️ {before - after} baris tanpa '{key_col}' dihapus ({clean_name})")

            if df.empty: continue

            # --- Standarkan header ---
            if clean_name in headers:
                df = df.reindex(columns=headers[clean_name])

            # --- Simpan sementara untuk clean ---
            combined_sheets.setdefault(clean_name, [])
            combined_sheets[clean_name].append(df)

            print(f"✅ {sheet_name} dari {brand} OK ({len(df)} baris)")
        except Exception as e:
            print(f"⚠️ Gagal baca sheet {sheet_name} dari {brand}: {e}")

# Gabungkan semua sheet
for name in combined_sheets:
    combined_sheets[name] = pd.concat(combined_sheets[name], ignore_index=True)

# ====================== CLEANING PROSES ======================
cleaned_sheets = {}

for sheet_name, df in combined_sheets.items():
    name_lower = sheet_name.lower()
    df_clean = df.copy()

    # --- SHOPEE ---
    if name_lower.startswith("shopee"):
        if "Waktu Pembayaran Dilakukan" in df_clean.columns:
            df_clean["Waktu Pembayaran Dilakukan"] = (
                df_clean["Waktu Pembayaran Dilakukan"]
                .apply(parse_datetime_safe)
                .dt.strftime("%Y-%m-%d %H:%M")
            )
        if "BRAND" in df_clean.columns:
            df_clean["BRAND"] = df_clean["BRAND"].apply(fix_brand)
        for c in ["Harga Sebelum Diskon","Harga Setelah Diskon","Jumlah","Total Harga Produk","Total Diskon"]:
            if c in df_clean.columns:
                df_clean[c] = clean_numeric(df_clean[c])

    # --- TOKPED ---
    elif name_lower.startswith("tokped"):
        if "Payment Date" in df_clean.columns:
            df_clean["Payment Date"] = (
                pd.to_datetime(df_clean["Payment Date"], errors="coerce")
                .dt.strftime("%d-%m-%Y %H:%M:%S")
            )
        if "BRAND" in df_clean.columns:
            df_clean["BRAND"] = df_clean["BRAND"].apply(fix_brand)

    # --- TIKTOK ---
    elif name_lower.startswith("tiktok"):
        if "Nama Brand" in df_clean.columns:
            df_clean["Nama Brand"] = df_clean["Nama Brand"].apply(fix_brand)

    # --- LAZADA ---
    elif name_lower.startswith("lazada"):
        if "Created at" in df_clean.columns:
            df_clean["Created at"] = (
                pd.to_datetime(df_clean["Created at"], errors="coerce")
                .dt.strftime("%m/%d/%Y")
            )
        for c in ["Paid Price", "Unit Price", "Shipping Fee"]:
            if c in df_clean.columns:
                df_clean[c] = clean_numeric(df_clean[c])

    cleaned_sheets[sheet_name] = df_clean

# ====================== SIMPAN DENGAN FORMAT ======================
print("\n💾 Membuat file akhir dengan format...")
with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
    for sheet_name, df in cleaned_sheets.items():
        df.to_excel(writer, sheet_name=sheet_name[:31], index=False)

# Copy format header (opsional, seperti clean.py)
wb_clean = load_workbook(output_file)
for ws in wb_clean.worksheets:
    for cell in ws[1]:
        cell.font = cell.font.copy(bold=True)
wb_clean.save(output_file)

print(f"\n✅ Merge + Cleaning selesai → {output_file}")
print(f"📊 Total sheets: {len(cleaned_sheets)}")
for s, df in cleaned_sheets.items():
    print(f"   - {s}: {len(df)} baris")
