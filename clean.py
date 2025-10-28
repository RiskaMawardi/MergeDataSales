import warnings
warnings.simplefilter("ignore", UserWarning)

import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

input_file = "eblo_clean.xlsx"
output_file = "EBLO.xlsx"

# ======================== CLEANING UTILS ========================
def clean_text_general(text: str) -> str:
    if pd.isna(text):
        return ""
    text = re.sub(r"[^a-zA-Z0-9\s.,*]", "", str(text))
    text = re.sub(r"\s+", " ", text).strip()
    return text

def clean_numeric(col_series):
    return (
        col_series.astype(str)
        .str.replace(r"[^0-9]", "", regex=True)
        .replace("", "0")
        .astype(float)
    )

def fix_brand(x):
    if pd.isna(x):
        return ""
    s = str(x).upper()
    original = s
    s = re.sub(r"^(TIKTOK|TOKOPEDIA)\s+", "", s)
    mapping = {
        "DR TEAL'S": "DR TEALS",
    }
    for key, val in mapping.items():
        if key in s:
            s = val
    print(f"[fix_brand] {original} -> {s}")   # Debug
    return s

def parse_datetime_safe(x):
    if pd.isna(x) or str(x).strip() == "":
        return ""
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M"):
        try:
            return pd.to_datetime(x, format=fmt)
        except ValueError:
            continue
    return pd.NaT
# ======================== CLEANING PROCESS ========================
# Baca semua sheet
all_sheets = pd.read_excel(input_file, sheet_name=None, engine="openpyxl")
cleaned_sheets = {}

for sheet_name, df in all_sheets.items():
    name_lower = sheet_name.lower()
    df_clean = df.copy()

    # ---------- SHOPEE ----------
    if name_lower.startswith("shopee"):
        if "Waktu Pembayaran Dilakukan" in df_clean.columns:
            df_clean["Waktu Pembayaran Dilakukan"] = (
                df_clean["Waktu Pembayaran Dilakukan"]
                .apply(parse_datetime_safe)
                .dt.strftime("%Y-%m-%d %H:%M")
            )
        if "Nomor Referensi SKU" in df_clean.columns and "SKU Induk" in df_clean.columns:
            df_clean["SKU Induk"] = df_clean["Nomor Referensi SKU"]

        if "Nama Produk" in df_clean.columns:
            df_clean["Nama Produk"] = df_clean["Nama Produk"].astype(str).str.slice(0, 100)

        if "BRAND" in df_clean.columns:
            df_clean["BRAND"] = df_clean["BRAND"].apply(fix_brand)

        for c in [
            "Harga Sebelum Diskon","Harga Setelah Diskon","Jumlah","Total Harga Produk",
            "Total Diskon","Diskon Dari Penjual","Diskon Dari Shopee"
        ]:
            if c in df_clean.columns:
                df_clean[c] = clean_numeric(df_clean[c])

        for c in ["Username (Pembeli)", "Nama Penerima", "No# Telepon", "Alamat Pengiriman"]:
            if c in df_clean.columns:
                df_clean[c] = df_clean[c].apply(clean_text_general)

    # ---------- TOKPED ----------
    elif name_lower.startswith("tokped"):
        if "Payment Date" in df_clean.columns:
            df_clean["Payment Date"] = (
                pd.to_datetime(df_clean["Payment Date"], errors="coerce")
                .dt.strftime("%d-%m-%Y %H:%M:%S")
            )
        if "Product Name" in df_clean.columns:
            df_clean["Product Name"] = df_clean["Product Name"].astype(str).str.slice(0, 100)

        if "BRAND" in df_clean.columns:
            df_clean["BRAND"] = df_clean["BRAND"].apply(fix_brand)

        if "Quantity" in df_clean.columns:
            df_clean["Quantity"] = (
                df_clean["Quantity"].astype(str)
                .str.replace(r"[^0-9]", "", regex=True)
                .replace("", "0")
                .astype(int)
            )
        for c in [
            "Price (Rp)", "Discount Amount (Rp)",
            "Subsidi Amount (Rp)", "Harga Jual (Rp)",
            "Total Amount (Rp)"
        ]:
            if c in df_clean.columns:
                df_clean[c] = clean_numeric(df_clean[c])
        for c in [
            "Shipping Price + fee (Rp)",
            "Insurance (Rp)",
            "Total Shipping Fee (Rp)"
        ]:
            if c in df_clean.columns:
                df_clean[c] = 0

    # ---------- TIKTOK ----------
    elif name_lower.startswith("tik tok"):
        if "Nama Brand" in df_clean.columns:
            df_clean["Nama Brand"] = df_clean["Nama Brand"].apply(fix_brand)

        if "Product Name" in df_clean.columns:
            df_clean["Product Name"] = df_clean["Product Name"].astype(str).str.slice(0, 100)

        if "Quantity" in df_clean.columns:
            df_clean["Quantity"] = (
                df_clean["Quantity"].astype(str)
                .str.replace(r"[^0-9]", "", regex=True)
                .replace("", "0")
                .astype(int)
            )

        if "Paid Time" in df_clean.columns:
            dt_parsed = pd.to_datetime(df_clean["Paid Time"], errors="coerce", dayfirst=True)
            df_clean["Paid Time"] = dt_parsed.dt.strftime("%m/%d/%Y")

        if "Recipient" in df_clean.columns:
            df_clean["Recipient"] = df_clean["Recipient"].apply(clean_text_general)

        if (
            "Order Amount" in df_clean.columns
            and "Total before platform Subsidy" in df_clean.columns
        ):
            df_clean["Total before platform Subsidy"] = clean_numeric(
                df_clean["Total before platform Subsidy"]
            )
            df_clean["Order Amount"] = df_clean["Total before platform Subsidy"]

    # ---------- LAZADA ----------
    elif name_lower.startswith("lazada"):
        if "Created at" in df_clean.columns:
            df_clean["Created at"] = (
                pd.to_datetime(df_clean["Created at"], errors="coerce")
                .dt.strftime("%m/%d/%Y")
            )

        if "Order Number" in df_clean.columns:
            for col_target in [
                "Shipping Phone Number",
                "Shipping Phone Number2",
                "Billing Phone Number",
                "Billing Phone Number2",
            ]:
                df_clean[col_target] = df_clean["Order Number"]

        for c in ["Paid Price", "Unit Price", "Shipping Fee"]:
            if c in df_clean.columns:
                df_clean[c] = clean_numeric(df_clean[c])

        if "Item Name" in df_clean.columns:
            df_clean["Item Name"] = df_clean["Item Name"].astype(str).str.slice(0, 100)

        if "BRAND" in df_clean.columns:
            df_clean["BRAND"] = df_clean["BRAND"].apply(fix_brand)

    cleaned_sheets[sheet_name] = df_clean

# ======================== COPY FORMAT & REPLACE DATA ========================
wb_orig = load_workbook(input_file)
wb_clean = load_workbook(input_file)  # copy workbook biar format ikut

for sheet_name, df_clean in cleaned_sheets.items():
    ws = wb_clean[sheet_name]

    # Hapus isi lama (tapi format tetap)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.value = None

    # Tulis ulang data bersih mulai dari baris ke-2
    for r_idx, row in enumerate(dataframe_to_rows(df_clean, index=False, header=False), start=2):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if sheet_name.lower().startswith("shopee"): 
                header_value = ws.cell(row=1, column=c_idx).value 
                if header_value in ["Harga Sebelum Diskon","Harga Setelah Diskon","Jumlah","Total Harga Produk",
            "Total Diskon","Diskon Dari Penjual","Diskon Dari Shopee","Perkiraan Ongkos Kirim","Total Pembayaran"]: 
                    cell.number_format = "General"
            if sheet_name.lower().startswith("lazada"): 
                header_value = ws.cell(row=1, column=c_idx).value 
                if header_value in ["Shipping Phone Number","Shipping Phone Number2","Billing Phone Number","Billing Phone Number2"]: 
                    cell.number_format = "@"
                    cell.value = "'" + str(value)
# Simpan hasil
wb_clean.save(output_file)

print(f"✅ Cleaning selesai dengan format ikut dari {input_file} → {output_file}")
